from __future__ import annotations

import csv
import json
import os
import secrets
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Tuple
from uuid import uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.models.ai_trading_manager import AiTmFile
from app.schemas.ai_files import AiFileMeta, AiFileSummary


DEFAULT_MAX_BYTES = 15 * 1024 * 1024  # 15MB


def _now() -> datetime:
    return datetime.now(UTC)


def _safe_filename(name: str) -> str:
    base = Path(name).name
    base = base.replace("\x00", "").strip()
    return base or "upload"


def _detect_kind(filename: str, content_type: str | None) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    if ext in {"csv"}:
        return "csv"
    if ext in {"xlsx"}:
        return "xlsx"
    if ext in {"png", "jpg", "jpeg", "webp", "gif"}:
        return "image"
    # Content-type fallbacks
    ct = (content_type or "").lower()
    if "csv" in ct:
        return "csv"
    if "spreadsheetml" in ct or "excel" in ct:
        return "xlsx"
    if ct.startswith("image/"):
        return "image"
    return "unknown"


def _upload_root(settings: Settings) -> Path:
    override = os.getenv("ST_AI_FILE_UPLOAD_DIR")
    if override:
        return Path(override)
    # Keep runtime artifacts inside backend/ by default.
    backend_root = Path(__file__).resolve().parents[3]
    return backend_root / "data" / "ai_uploads"


def _ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _copy_to_disk(upload: UploadFile, dst: Path, *, max_bytes: int) -> int:
    total = 0
    with dst.open("wb") as f:
        while True:
            chunk = upload.file.read(1024 * 1024)
            if not chunk:
                break
            total += len(chunk)
            if total > max_bytes:
                raise HTTPException(
                    status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    detail=f"File too large (max {max_bytes} bytes).",
                )
            f.write(chunk)
    return total


def _csv_summary(path: Path) -> AiFileSummary:
    columns: list[str] = []
    preview: list[dict[str, Any]] = []
    row_count = 0

    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return AiFileSummary(kind="csv", columns=[], row_count=0, preview_rows=[])

        columns = [str(c or "").strip() for c in header]
        # Ensure non-empty column names.
        for i, c in enumerate(columns):
            if c:
                continue
            columns[i] = f"col_{i+1}"

        for row in reader:
            row_count += 1
            if len(preview) < 5:
                obj = {columns[i]: (row[i] if i < len(row) else "") for i in range(len(columns))}
                preview.append(obj)

    return AiFileSummary(kind="csv", columns=columns, row_count=row_count, preview_rows=preview)


def _xlsx_summary(path: Path) -> AiFileSummary:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheets = list(wb.sheetnames or [])
    if not sheets:
        return AiFileSummary(kind="xlsx", sheets=[], active_sheet=None, columns=[], row_count=0, preview_rows=[])

    sheet_name = sheets[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return AiFileSummary(
            kind="xlsx",
            sheets=sheets,
            active_sheet=sheet_name,
            columns=[],
            row_count=0,
            preview_rows=[],
        )

    raw_cols = list(header_row or [])
    columns: list[str] = []
    for i, v in enumerate(raw_cols):
        s = str(v or "").strip()
        columns.append(s or f"col_{i+1}")

    preview: list[dict[str, Any]] = []
    row_count = 0
    for r in rows_iter:
        row_count += 1
        if len(preview) < 5:
            cells = list(r or [])
            obj = {}
            for i, col in enumerate(columns):
                val = cells[i] if i < len(cells) else None
                # JSON-safe value
                if isinstance(val, (datetime,)):
                    obj[col] = val.isoformat()
                else:
                    obj[col] = val
            preview.append(obj)

    return AiFileSummary(
        kind="xlsx",
        sheets=sheets,
        active_sheet=sheet_name,
        columns=columns,
        row_count=row_count,
        preview_rows=preview,
    )


def summarize_file(kind: str, path: Path) -> AiFileSummary:
    if kind == "csv":
        return _csv_summary(path)
    if kind == "xlsx":
        return _xlsx_summary(path)
    return AiFileSummary(kind=kind, columns=[], row_count=0, preview_rows=[])


def _serialize_summary(summary: AiFileSummary) -> str:
    return json.dumps(summary.model_dump(mode="json"), ensure_ascii=False, sort_keys=True, default=str)


def _parse_summary(raw: str) -> AiFileSummary:
    try:
        parsed = json.loads(raw) if raw else {}
    except Exception:
        parsed = {}
    if not isinstance(parsed, dict):
        parsed = {}
    return AiFileSummary.model_validate(parsed)


def create_file(
    db: Session,
    settings: Settings,
    *,
    user_id: int,
    upload: UploadFile,
    max_bytes: int = DEFAULT_MAX_BYTES,
) -> AiFileMeta:
    filename = _safe_filename(upload.filename or "upload")
    kind = _detect_kind(filename, upload.content_type)
    if kind not in {"csv", "xlsx", "image"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Only .csv, .xlsx, and images are supported.",
        )

    root = _upload_root(settings) / f"user_{user_id}"
    _ensure_dir(root)

    file_id = uuid4().hex
    # Preserve extension for convenience.
    ext = Path(filename).suffix.lower() or (".csv" if kind == "csv" else ".xlsx" if kind == "xlsx" else ".png")
    storage_name = f"{file_id}{ext}"
    path = root / storage_name

    # Best-effort: ensure we don't overwrite.
    if path.exists():
        path = root / f"{file_id}-{secrets.token_hex(4)}{ext}"

    size = _copy_to_disk(upload, path, max_bytes=max_bytes)
    summary = summarize_file(kind, path)

    row = AiTmFile(
        file_id=file_id,
        user_id=user_id,
        filename=filename,
        content_type=upload.content_type,
        size_bytes=int(size),
        storage_path=str(path),
        summary_json=_serialize_summary(summary),
        created_at=_now(),
        updated_at=_now(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)

    return AiFileMeta(
        file_id=row.file_id,
        filename=row.filename,
        size=int(row.size_bytes or 0),
        mime=row.content_type,
        created_at=row.created_at,
        summary=_parse_summary(row.summary_json),
    )


def get_file_meta(
    db: Session,
    *,
    file_id: str,
    user_id: int,
) -> AiFileMeta | None:
    row = (
        db.query(AiTmFile)
        .filter(AiTmFile.file_id == file_id, AiTmFile.user_id == user_id)
        .one_or_none()
    )
    if row is None:
        return None
    return AiFileMeta(
        file_id=row.file_id,
        filename=row.filename,
        size=int(row.size_bytes or 0),
        mime=row.content_type,
        created_at=row.created_at,
        summary=_parse_summary(row.summary_json),
    )


def get_file_path(
    db: Session,
    *,
    file_id: str,
    user_id: int,
) -> Tuple[str, Path] | None:
    row = (
        db.query(AiTmFile)
        .filter(AiTmFile.file_id == file_id, AiTmFile.user_id == user_id)
        .one_or_none()
    )
    if row is None:
        return None
    return row.filename, Path(row.storage_path)


__all__ = ["DEFAULT_MAX_BYTES", "create_file", "get_file_meta", "get_file_path"]
